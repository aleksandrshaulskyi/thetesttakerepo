from asyncio import get_event_loop
from pathlib import Path
from typing import Any, AsyncGenerator, Iterable

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from settings import settings

from application.exceptions import OpenPyXlException
from application.ports import ExcelReaderPort
from infrastructure.utils import product_data_builder


class ExcelReader(ExcelReaderPort):
    '''
    The actual implementation of the excel reader port.
    '''

    def execute(self, path: str, sheet: str, batch_size: int) -> Iterable[list[tuple[Any, ...]]]:
        try:
            workbook = load_workbook(filename=path, data_only=True, read_only=True, keep_links=False)
            print(f'Opened the workbook {workbook}')
        except FileNotFoundError:
            raise OpenPyXlException('A file with provided path does not exist.')
        except InvalidFileException:
            raise OpenPyXlException('Tried to open invalid file.')
        else:
            try:
                worksheet = workbook[sheet]
                print(f'Opened the worksheet {worksheet}')
            except KeyError:
                raise OpenPyXlException(f'There is no {sheet} in the opened workbook.')
            else:
                print('Reading table in batches...')
                try:
                    batch = []

                    for row in worksheet.iter_rows(
                        values_only=True,
                        max_row=settings.excel_max_row,
                        min_row=3,
                        max_col=settings.excel_max_col,
                    ):
                        if any(cell is None for cell in row):
                            print('Found a row containing None values, ignoring it...')
                            continue
                        batch.append(row)
                        if len(batch) >= batch_size:
                            batch = self.transform_batch(batch=batch, path=path)
                            yield batch
                            batch = []
                    if batch:
                        batch = self.transform_batch(batch=batch, path=path)
                        yield batch
                finally:
                    workbook.close()
                    print('Finished reading...')

    def transform_batch(self, batch: list[tuple], path) -> list[dict]:
        file_name = path.split('/')[-1]
        transformed_batch = []
        for item in batch:
            transformed_batch.append(product_data_builder.build(item, file_name))
        return transformed_batch

    async def read(self, path: str, sheet: str, batch_size: int) -> AsyncGenerator[list[tuple[Any, ...]], None]:
        event_loop = get_event_loop()
        iterable = iter(self.execute(path, sheet, batch_size))

        while True:
            batch = await event_loop.run_in_executor(None, lambda: next(iterable, None))
            if batch is None:
                Path(path).unlink(missing_ok=True)
                break
            yield batch
