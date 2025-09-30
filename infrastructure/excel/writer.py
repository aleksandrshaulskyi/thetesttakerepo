from asyncio import get_event_loop
from os.path import exists

from openpyxl import load_workbook, Workbook


class ExcelWriter:

    def execute(self, rows_data: list, path: str) -> None:
        if exists(path=path):
            workbook = load_workbook(path)
        else:
            workbook = Workbook()

        worksheet = workbook.active

        for row_data in rows_data:
            data_to_write = list(row_data.values())
            worksheet.append(data_to_write)

        workbook.save(path)

    async def write_rows(self, rows_data: list, path: str) -> None:
        event_loop = get_event_loop()

        await event_loop.run_in_executor(None, self.execute, rows_data, path)
